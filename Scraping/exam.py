from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import tkinter
import tkinter.messagebox as msgbox
import csv
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import chromedriver_autoinstaller


#Tkクラスの作成
root = tkinter.Tk()
#画面サイズ
root.geometry('300x200')
#画面タイトル
root.title('受けたい試験を入力')

#テキストボックスの作成
lbl_1 = tkinter.Label(text='アルファベット大文字2文字')
lbl_1.place(x=30, y=30)
txt_1 = tkinter.Entry(width=20)
txt_1.place(x=30,y=50)
lbl_2 = tkinter.Label(text='3桁の数字')
lbl_2.place(x=30, y=70)
txt_2 = tkinter.Entry(width=20)
txt_2.place(x=30,y=90)

#examをグローバル変数として宣言
exam=str("")

#ボタンのクリックイベント
def btnclick():
    global exam
    exam=str(txt_1.get())+"-"+str(txt_2.get())
    msgbox.showinfo('送信完了', exam+'送信が完了しました')
    root.destroy()  

#ボタンの作成
btn = tkinter.Button(root, text='送信', command=btnclick)
btn.place(x=140,y=170)
#表示
root.mainloop()

# ChromeDriverを自動的にダウンロードおよび更新
chromedriver_autoinstaller.install()
driver = webdriver.Chrome()

# # chromedriver.exeがある場所
# driver_path = "./chromedriver.exe"

# # ドライバーを最小化した状態で動かすオプションの作成（処理の問題なのかちゃんとURLが取得できなくなる）
# #--------------------------------------------------------------
# # options=Options()
# # options.add_argument('--headless')
# # Serviceクラスにドライバーのパスを設定して、Serviceをwebdriverに与えてdriverを作る
# #--------------------------------------------------------------
# service = Service(executable_path=driver_path)
# driver = webdriver.Chrome(service=service,)

# examtopicsのURL
url = 'https://www.examtopics.com/discussions/microsoft'
#CSVファイルを開く（なければ新規作成）
with open('mcp/csv/'+exam+'.csv', 'w', newline='') as file:
    writer = csv.writer(file)

    # urlが続く間繰り返す
    while(url != ''):

    # ページを開く
        driver.get(url)
     
        # 要素を取得（ExamtopicのHTMLで試験のリンクがdiscussion-linkであるためここで指定）
        elements = driver.find_elements(By.CLASS_NAME, 'discussion-link')

        # 取得した要素を処理する
        for element in elements:

            # 取得した要素から、特定の試験を含むものを抽出
            # ※ここで取得したい試験のキーワードを指定する
            if (exam in element.text):
                url = element.get_attribute('href')
                print(url)
                writer.writerow([url])


    # 次ページがあるか確認
        elements_btn = driver.find_elements(By.CLASS_NAME, 'btn-sm')

    # 取得した要素を処理する（次ページのURLを探す）
        url = ''
        for element_btn in elements_btn:
         if ('Next' in element_btn.text):
                url = element_btn.get_attribute('href')
                break

# ブラウザを閉じる。
driver.quit()

print('csvに書き込みました')

# CSVファイルを読み込む
df = pd.read_csv('mcp/csv/'+exam+'.csv')

# Excelファイルに書き込む(A2から)
with pd.ExcelWriter('mcp/xlsx/'+exam+'.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, index=False,startrow=1)

# Excelファイルを開いてハイパーリンクに変更

wb = load_workbook('mcp/xlsx/'+exam+'.xlsx')
ws = wb.active

ws['A1'] = '問題リンク'

for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
    for cell in row:
        cell.hyperlink = cell.value
        cell.font = Font(color="0000EE", underline="single")

wb.save('mcp/xlsx/'+exam+'.xlsx')

print("Excelに変換しました")
